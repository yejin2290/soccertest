import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import win32com.client
import os
import poisson_distribution

class Make_Excel():

    # 팀 이름을 전달해주어서 history출력시키고, 임의로 스탯 설정한 것도 받아서 출력시켜야함.
    def __init__(self, TEAM1, TEAM2, Expected_result):
        self.TEAM1 = TEAM1
        self.TEAM2 = TEAM2
        self.Expected_result = Expected_result

    def start(self):
        self.TEAM1 = self.TEAM1.replace(" ", "-")
        self.TEAM2 = self.TEAM2.replace("-", " ")
        print("---------------")
        print(self.TEAM1)
        print(self.TEAM2)
        url = 'https://www.11v11.com/teams/' + self.TEAM1 + '/tab/opposingTeams/opposition/' + self.TEAM2 + '/'

        # 엑셀 파일에 쓰기
        write_wb = Workbook()
        write_ws = write_wb.active

        response = requests.get(url)
        soup = BeautifulSoup(response.content, 'lxml')

        # 홈페이지 엤는 table을 찾는 과정. 저 table 이름이 홈페이지에 있는 선수들 스탯 창
        table = soup.find("table",
                          {"class": "sortable"})
        data = []
        for a in table.find_all("tr"):
            for b in a.find_all("td"):
                data.append(b.get_text())
        i = 0
        result_list = []
        result = []
        count_play = 0
        count_win = 0
        count_lose = 0
        count_draw = 0
        team1_score = 0
        team2_score = 0
        for i in range(0, int(len(data) / 5)):

            # 2015 이전의 자료들을 제외하기 위한 조건문
            if (int(data[5 * i].split(" ")[2]) < 2015):
                continue
            else:
                count_play = count_play + 1
                # 사이트에서 뒤죽박죽으로 결과가 저장되어 있어 TEAM1 v TEAM2로 결과를 재정렬함.
                if data[5 * i + 1][0] == self.TEAM2[0].capitalize():
                    data[5 * i + 1] = self.TEAM1.title() + " v " + self.TEAM2.title()
                    j = data[5 * i + 3].split("-")
                    team1_score = team1_score + int(j[1])
                    team2_score = team2_score + int(j[0])
                    data[5 * i + 3] = j[1] + "-" + j[0]
                else:
                    j = data[5 * i + 3].split("-")
                    team1_score = team1_score + int(j[0])
                    team2_score = team2_score + int(j[1])

                # 경기 승률 체크
                if data[5 * i + 2] == "W":
                    count_win = count_win + 1
                elif data[5 * i + 2] == "L":
                    count_lose = count_lose + 1
                else:
                    count_draw = count_draw + 1

                # result 리스트에 필요한 정보만을 저장함.
                result.append(data[5 * i + 0])
                result.append(data[5 * i + 2])
                result.append(data[5 * i + 3])

                result_list.append(result)
                result = []
        # result_list.append([count_play, count_win, count_lose, count_draw])
        for i in result_list:
            print(i)
        print("play: %d win: %d lose %d draw: %d" % (count_play, count_win, count_lose, count_draw))
        print("team1_score_aver: %.3f team2_score_aver: %.3f" % (team1_score / count_play, team2_score / count_play))
        print(os.getcwd())
        write_ws['A1'] = self.TEAM1 + "  vs  " + self.TEAM2
        write_ws.append(["Date", "Result", "Score", "", "Average score", "", "Play", "Win", "Lose", "Draw"])

        for i in result_list:
            write_ws.append(i)

        # 엑셀 작성
        write_ws['E1'] = "<History>"
        write_ws['E5'] = "<Average Win>"
        write_ws['E3'] = str(format(team1_score / count_play, ".3f")) + " vs " + str(
            format(team2_score / count_play, ".3f"))
        write_ws['E6'] = format(count_win / count_play * 100, ".3f") + "%"
        write_ws['G3'] = (count_play)
        write_ws['H3'] = (count_win)
        write_ws['I3'] = (count_lose)
        write_ws['J3'] = (count_draw)
        write_ws['E9'] = "<MY RESULT>"
        write_ws['L1'] = "<POISSON>"
        write_ws['M2'] = self.TEAM1
        write_ws['M3'] = self.TEAM2
        for i in range(0, 10):
            write_ws.cell(row = 1, column = i + 14).value = i
        # poisson에 해당 팀들의 평균 골의 값을 넘긴 뒤, 스코어 중 확률이 가장 높은 것 엑셀에 작성
        hel = poisson_distribution.Poisson(team1_score / count_play, team2_score / count_play)
        [list_1, list_2] = hel.get_poisson_list()

        # 엑셀에 team1과 team2의 poisson 값들을 작성. 아스키와 문자를 이용하여 엑셀에 차례대로 입력시킴
        for i in range (2, 12):
            write_ws.cell(row=2, column= 14 + (i - 2)).value = (format(float(list_1[i - 2][1] * 100)  , ".3f")) + "%"
            write_ws.cell(row=3, column= 14 + (i - 2)).value = (format(float(list_2[i - 2][1] * 100)  , ".3f")) + "%"

        # poisson에서 최대 확률을 찾음으로써 스코어 결과 도출
        max_1 = 0.0
        max_2 = 0.0

        for i in range (0, 10):
            if max_1 <= list_1[i][1] :
                max_1 = list_1[i][1]
                max_1_index = i
            if max_2 <= list_2[i][1] :
                max_2 = list_2[i][1]
                max_2_index = i
        # poisson으로 구한 결과 엑셀에 표시. 결과는 노란색으로 입히기
        write_ws['L5'] = "<POISSON RESULT>"
        yellowFill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')
        write_ws.cell(row=6, column=12).fill = yellowFill

        if max_1_index > max_2_index :
            write_ws['L6'] = str(max_1_index) + " vs " + str(max_2_index) + " (W) "
        elif max_1_index == max_2_index:
            write_ws['L6'] = str(max_1_index) + " vs " + str(max_2_index) + " (D) "
        else:
            write_ws['L6'] = str(max_1_index) + " vs " + str(max_2_index) + " (L) "

        # 엑셀에 예상된 결과를 출력시킴, 색깔도 입히기
        # FFFF99
        redFill = PatternFill(start_color='FF4500', end_color='FF4500', fill_type='solid')
        GreenFill = PatternFill(start_color='008000', end_color='008000', fill_type='solid')
        write_ws['E6'].fill = GreenFill
        for i in range(10, 41):
            write_ws.cell(row = i, column= 5).fill = redFill
            if int(self.Expected_result["goal"][0]) > int(self.Expected_result["goal"][1]):
                write_ws.cell(row=i, column=5).value = self.Expected_result["goal"][0] + " vs " + \
                                                       self.Expected_result["goal"][1] + "(W)"
            elif int(self.Expected_result["goal"][0]) == int(self.Expected_result["goal"][1]):
                write_ws.cell(row=i, column=5).value = self.Expected_result["goal"][0] + " vs " + \
                                                       self.Expected_result["goal"][1] + "(D)"
            else:
                write_ws.cell(row= i, column= 5).value = self.Expected_result["goal"][0] + " vs " + \
                                                         self.Expected_result["goal"][1] + "(L)"
            break
        write_wb.save(os.getcwd() + "/result_list.xlsx")
        write_wb.close()

        # 엑셀 파일 open
        excel = win32com.client.Dispatch("Excel.Application")
        workbook = excel.Workbooks.Open(os.getcwd() + "/result_list.xlsx")

        return excel



#print(hell)